import operator
import re
import pandas as pd
import openpyxl as xl
import json
import matplotlib
import matplotlib.pyplot as plt

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape

from reportlab.lib.enums import TA_JUSTIFY
from reportlab.platypus import SimpleDocTemplate, Image, Table, Paragraph, Frame, PageTemplate
from reportlab.platypus.tableofcontents import TableOfContents, SimpleIndex
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import utils
from reportlab.lib import colors

rx_dict = {
    'setTable': r"\$setTable (#path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+) #colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #rowRange (?P<rowRange>[0-9]+:[0-9]+))+ as (?P<var>([a-zA-Z]|[0-9]|-|_)+)",
    'loopContent' : r"^_(?P<content>(.*))",
    'ifContent' : r"^@(?P<content>(.*))",
    'insertArtifact': r"\$insert(?P<type>Artifact) #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+) #colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #rowRange (?P<rowRange>[0-9]+:[0-9]+)",
    'insertArtifact2': r"\$insert(?P<type>Artifact) (?P<var>([a-zA-Z]|[0-9]|-|_)+)",
    'insertCompoundArtifact': r"\$insert(?P<type>Artifact) #key (?P<key>([a-zA-Z]|[0-9]|-|_| )+)? (<#path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+) #colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #rowRange (?P<rowRange>[0-9]+:[0-9]+)>) (<#path2 (?P<path2>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet2 (?P<sheet2>([a-zA-Z]|[0-9]|-|_| )+) #colRange2 (?P<colRange2>[a-zA-Z]+:[a-zA-Z]+) #rowRange2 (?P<rowRange2>[0-9]+:[0-9]+)>)",
    'insertCompoundArtifact2': r"\$insert(?P<type>Artifact) #key (?P<key>([a-zA-Z]|[0-9]|-|_| )+)? (<(?P<var>([a-zA-Z]|[0-9]|-|_)+)>) (<(?P<var2>([a-zA-Z]|[0-9]|-|_)+)>)",
    'insertImage': r"\$insert(?P<type>Image) #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) (#width (?P<width>([0-9]+)))?",
    'insertText':r"\$insertText (#text <(?P<text>([a-zA-Z]|[0-9]|-|_| |\.|,)+)>)",
    'insertHistogram':r"\$insertHistogram #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+) #colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #rowRange (?P<rowRange>[0-9]+:[0-9]+) #xAxis \[(?P<xAxis>((([a-zA-Z]|[0-9]|-|_| )+),?)+)\] #yAxis \[(?P<yAxis>((([a-zA-Z]|[0-9]|-|_| )+),?)+)\]",
    'insertPieChart':r"\$insertPieChart #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) #sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+) #colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #rowRange (?P<rowRange>[0-9]+:[0-9]+) #data \[(?P<data>((([a-zA-Z]|[0-9]|-|_| )+),?)+)\] #label \[(?P<label>((([a-zA-Z]|[0-9]|-|_| )+),?)+)\]",
    'loopRow' : r"\$forEach (?P<loopType>(row|col|sheet)) in #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) (#sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+))? (#rowRange (?P<rowRange>[0-9]+:[0-9]+) #col (?P<col>([a-zA-Z]|[0-9]|-|_| )+))?",
    'loopSheet' : r"\$forEach (?P<loopType>(row|col|sheet)) in #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+)",
    'loopCol' : r"\$forEach (?P<loopType>(row|col|sheet)) in #path (?P<path>([a-zA-Z]|[0-9]|-|_| |\.)+) (#sheet (?P<sheet>([a-zA-Z]|[0-9]|-|_| )+))? (#colRange (?P<colRange>[a-zA-Z]+:[a-zA-Z]+) #row (?P<row>[0-9]+))?",
    'loopEnd' : r"\$endForEach",
    'ifHeader' : r"\$if (?P<first>(([a-zA-Z]|[0-9]|-|_|\.)+)) (?P<logic>(==|!=|<=|>=|<|>)) (?P<second>(([a-zA-Z]|[0-9]|-|_|\.)+))",
    'ifEnd' : r"\$endIf",
    'createChapter' : r"\$createChapter ((#text <(?P<text>([a-zA-Z]|[0-9]|-|_| |\.)+)>)|(?P<row>(row))|(?P<col>(col)))",
    'createSection' : r"\$createSection ((#text <(?P<text>([a-zA-Z]|[0-9]|-|_| |\.)+)>)|(?P<row>(row))|(?P<col>(col)))",
}

def getFrame(self,framename,orientation="Portrait"):
    """
    returns frame
    frame._x1,frame._y1
    frame._width,frame._height
    frame._leftPadding,frame._bottomPadding
    frame._rightPadding,frame._topPadding
    """

    f = operator.attrgetter("id")
    frame = None

    for temp in self.pageTemplates[::-1]:
        if f(temp) == framename:
            #print( temp.id )
            for frame in temp.frames:
                print( frame.id )
                if f(frame) == orientation:

                    return frame

def _parse_line(line):
    
    # Do a REGEX search for all defined regexes and return the key and match result of the first matching regex    
    for key, rx in rx_dict.items():
        match = re.search(rx, line)
        if match:
            return key, match
        
    # If there are no matches
    return None, None

def parse_file(filepath, story):
    
    # Create an empty list to collect all the data
    data = []   
    dataFrames = {}
    
    # Open the file and read line by line
    with open (filepath, "r") as file_object:
        line = file_object.readline()
        
        loopType = "" 
        loopPath = ""
        loopSheet = ""   
        loopRowRange = ""  
        loopColRange = ""
        loopRow = ""
        loopCol = ""
        loopContent = []
        
        ifFirst = ""
        ifSecond = ""
        ifLogic= ""
        ifContent = []
        
        
        while line:
                        
            # At each line check for a match with a regex
            key, match = _parse_line(line)
            
            if key == 'setTable':
                skipRows = int(re.search(r'(\d+):(\d+)', match.group('rowRange')).group(1)) - 1
                df = pd.read_excel('.\Artifacts\%s' % (match.group('path')), match.group('sheet'), skiprows = skipRows, header=0, usecols=match.group('colRange'))
                varName = match.group('var')
                if varName not in dataFrames:
                    dataFrames.update({varName:df})    
                    
            if key == 'insertArtifact2':
                _insert_artifact2(dataFrames[match.group('var')], story)
            
            # Extract insertType
            if key == 'insertArtifact':
                type = match.group('type')
                path = match.group('path')
                sheet = match.group('sheet')
                colRange = match.group('colRange')
                rowRange = match.group('rowRange')
                skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
                rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
                _insert_artifact(path, sheet, colRange, skipRows, story)
                                
                row = {
                        'Type': type,
                        'Path': path,
                    }
                
                data.append(row)
                
            if key == 'insertCompoundArtifact':
                print("Reconoce")
                keyArtifact = match.group('key')
                path = match.group('path')
                sheet = match.group('sheet')
                colRange = match.group('colRange')
                rowRange = match.group('rowRange')
                skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
                rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
                path2 = match.group('path2')
                sheet2 = match.group('sheet2')
                colRange2 = match.group('colRange2')
                rowRange2 = match.group('rowRange2')
                skipRows2 = int(re.search(r'(\d+):(\d+)', rowRange2).group(1)) - 1
                rowNumber2 = int(re.search(r'(\d+):(\d+)', rowRange2).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange2).group(1))
                _insert_compound_artifact(keyArtifact, path, sheet, colRange, skipRows, path2, sheet2, colRange2, skipRows2, story)
                
            if key == 'insertCompoundArtifact2':
                keyArtifact = match.group('key')
                varName = match.group('var')
                varName2 = match.group('var2')
                _insert_compound_artifact2(keyArtifact, dataFrames[varName], dataFrames[varName2], story)
                
            if key == 'insertImage':
                type = match.group('type')
                path = match.group('path')
                width = int(match.group('width'))
                print(width)
                _insert_image(path, width, story)
                                
                row = {
                        'Type': type,
                        'Path': path,
                    }
                
                data.append(row)
                
            if key == 'insertText':
                textInput = match.group('text')
                _insert_text(textInput)
                
            if key == 'insertHistogram':
                path = match.group('path')
                sheet = match.group('sheet')
                colRange = match.group('colRange')
                rowRange = match.group('rowRange')
                skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
                rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
                xAxis = match.group('xAxis')
                yAxis = match.group('yAxis')

                _insert_histogram(path, sheet, colRange, skipRows, xAxis, yAxis, story)
                
            if key == 'insertPieChart':
                path = match.group('path')
                sheet = match.group('sheet')
                colRange = match.group('colRange')
                rowRange = match.group('rowRange')
                skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
                rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
                data = match.group('data')
                label = match.group('label')

                _insert_piechart(path, sheet, colRange, skipRows, data, label, story)
                
            if key == 'loopRow':
                loopType = match.group('loopType')
                loopPath = match.group('path')
                if (loopType == 'row'):
                    loopSheet = match.group('sheet')
                    loopRowRange = match.group('rowRange')
                    loopCol = match.group('col')
                    
            if key == 'loopSheet':
                loopType = match.group('loopType')
                loopPath = match.group('path')
                    
            if key == 'loopCol':
                loopType = match.group('loopType')
                loopPath = match.group('path')
                loopSheet = match.group('sheet')
                loopColRange = match.group('colRange')
                loopRow = match.group('row')
                    
            if key == 'loopContent':
                loopContent.append(match.group('content'))
                
            if key == 'loopEnd':
                if loopType == "row":
                    _loop_row(loopPath, loopSheet, loopRowRange, loopCol, loopContent, story)
                    loopContent = []
                if loopType == "sheet":
                    _loop_sheet(loopPath, loopContent, story)
                    loopContent = []
                if loopType == "col":
                    _loop_col(loopPath, loopSheet, loopColRange, loopRow, loopContent, story)
                    loopContent = []
                    
            if key == 'createChapter':
                textInput = match.group('text')
                _create_chapter(textInput)
                
            if key == 'createSection':
                textInput = match.group('text')
                _create_section(textInput)
                
            if key == 'ifHeader':
                ifFirst = match.group('first')
                ifSecond = match.group('second')
                ifLogic = match.group('logic')
                
            if key == 'ifContent':
                ifContent.append(match.group('content'))
                
            if key == 'ifEnd':
                _conditional(ifFirst, ifSecond, ifLogic, ifContent, story)
                ifContent = []
                               
            line = file_object.readline()
            
def parse_string_array(content, story, **kwargs):
    
    print("Llega al string array y el value es: "+kwargs.get('value','cell.value'))
    
    # Create an empty list to collect all the data
    data = []   
        
    loopType = "" 
    loopPath = ""
    loopSheet = ""   
    loopRowRange = ""  
    loopColRange = ""
    loopRow = ""
    loopCol = ""
    loopContent = []
    
    ifFirst = ""
    ifSecond = ""
    ifLogic= ""
    ifContent = []
        
    for line in content:
                    
        # At each line check for a match with a regex
        key, match = _parse_line(line)
        
        # Extract insertType
        if key == 'insertArtifact':
            type = match.group('type')
            path = match.group('path')
            sheet = match.group('sheet')
            print(sheet)
            if sheet == "sheetName":
                print("Entró acá")
                sheet = kwargs.get('value', 'sheetName')
            colRange = match.group('colRange')
            rowRange = match.group('rowRange')
            skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
            rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
            _insert_artifact(path, sheet, colRange, skipRows, story)
                        
            row = {
                    'Type': type,
                    'Path': path,
                }
            
            data.append(row)
            
        if key == 'insertCompoundArtifact':
            keyArtifact = match.group('key')
            path1 = match.group('path1')
            sheet1 = match.group('sheet1')
            if sheet1 == "sheetName":
                sheet1 = kwargs.get('value', 'sheetName')
            colRange1 = match.group('colRange1')
            rowRange1 = match.group('rowRange1')
            skipRows1 = int(re.search(r'(\d+):(\d+)', rowRange1).group(1)) - 1
            rowNumber1 = int(re.search(r'(\d+):(\d+)', rowRange1).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange1).group(1))
            path2 = match.group('path1')
            sheet2 = match.group('sheet1')
            if sheet2 == "sheetName":
                sheet2 = kwargs.get('value', 'sheetName')
            colRange2 = match.group('colRange1')
            rowRange2 = match.group('rowRange1')
            skipRows2 = int(re.search(r'(\d+):(\d+)', rowRange2).group(1)) - 1
            rowNumber2 = int(re.search(r'(\d+):(\d+)', rowRange2).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange2).group(1))
            _insert_compound_artifact(keyArtifact, path1, sheet1, colRange1, skipRows1, path2, sheet2, colRange2, skipRows2, story)
            
        if key == 'insertImage':
            type = match.group('type')
            path = match.group('path')
            width = int(match.group('width'))
            _insert_image(path, width, story)
                        
            row = {
                    'Type': type,
                    'Path': path,
                }
            
            data.append(row)
            
        if key == 'insertText':
            textInput = match.group('text')
            _insert_text(textInput)
            
        if key == 'insertHistogram':
            path = match.group('path')
            sheet = match.group('sheet')
            if sheet == "sheetName":
                sheet = kwargs.get('value', 'sheetName')
            colRange = match.group('colRange')
            rowRange = match.group('rowRange')
            skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
            rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
            xAxis = match.group('xAxis')
            yAxis = match.group('yAxis')

            _insert_histogram(path, sheet, colRange, skipRows, xAxis, yAxis, story)
                
        if key == 'insertPieChart':
            path = match.group('path')
            sheet = match.group('sheet')
            if sheet == "sheetName":
                sheet = kwargs.get('value', 'sheetName')
            colRange = match.group('colRange')
            rowRange = match.group('rowRange')
            skipRows = int(re.search(r'(\d+):(\d+)', rowRange).group(1)) - 1
            rowNumber = int(re.search(r'(\d+):(\d+)', rowRange).group(2)) - int(re.search(r'(\d+):(\d+)', rowRange).group(1))
            data = match.group('data')
            label = match.group('label')

            _insert_piechart(path, sheet, colRange, skipRows, data, label, story)
            
        if key == 'loopRow':
            loopType = match.group('loopType')
            loopPath = match.group('path')
            if (loopType == 'row'):
                loopSheet = match.group('sheet')
                if loopSheet == "sheetName":
                    loopSheet = kwargs.get('value', 'sheetName')
                loopRowRange = match.group('rowRange')
                loopCol = match.group('col')
                
        if key == 'loopSheet':
                loopType = match.group('loopType')
                loopPath = match.group('path')
                
        if key == 'loopCol':
                loopType = match.group('loopType')
                loopPath = match.group('path')
                loopSheet = match.group('sheet')
                if loopSheet == "sheetName":
                    loopSheet = kwargs.get('value', 'sheetName')
                loopColRange = match.group('colRange')
                loopRow = match.group('row')
                
        if key == 'loopContent':
            loopContent.append(match.group('content'))
            
        if key == 'loopEnd':
            if loopType == "row":
                _loop_row(loopPath, loopSheet, loopRowRange, loopCol, loopContent, story)
                loopContent = []
            if loopType == "sheet":
                _loop_sheet(loopPath, loopContent, story)
                loopContent = []
            if loopType == "col":
                _loop_col(loopPath, loopSheet, loopColRange, loopRow, loopContent, story)
                loopContent = []
                
        if key == 'createChapter':
            print("Va a crear capítulo en el parse string array")
            textInput = match.group('text')
            if textInput == "cell.value":
                textInput = kwargs.get('value', 'cell.value')
            _create_chapter(textInput)
            
        if key == 'createSection':
            textInput = match.group('text')
            if textInput == "cell.value":
                textInput = kwargs.get('value', 'cell.value')
            _create_section(textInput)
            
        if key == 'ifHeader':
            ifFirst = match.group('first')
            ifSecond = match.group('second')
            ifLogic = match.group('logic')
                
        if key == 'ifContent':
            ifContent.append(match.group('content'))
                
        if key == 'ifEnd':
            if (ifFirst == 'cell.value'):
                ifFirst = kwargs.get('value', 'cell.value')
            elif (ifSecond == 'cell.value'):
                ifSecond = kwargs.get('value', 'cell.value')
            _conditional(ifFirst, ifSecond, ifLogic, ifContent, story, value = kwargs.get('value', 'cell.value'))
            ifContent = []      
    
def _insert_artifact(path, sheet, colRange, skipRows, Story):
    print("Path: % s" % (path))
    print("Sheet: % s" % (sheet))
    print("Column Range: % s" % (colRange))
    print("Skip this row number: % s" % (skipRows))
    
    # To open Workbook
    wb = xl.load_workbook(filename = '.\Artifacts\%s' % (path))
    # Get the active worksheet
    ws = wb.active 
    # Print the artifact's name
    print("The Artifact Title is: %s" % (ws['A3'].value))
    
    # Save the data into a DataFrame from Pandas
    df = pd.read_excel('.\Artifacts\%s' % (path), sheet, skiprows = skipRows, header=0, usecols=colRange)
    
    print(df)
    
    # Transform the dataframe to a list in order to be printed in the document
    lista = [df.columns[:,].values.astype(str).tolist()] + df.values.tolist()
    
    newList = []
    
    print(lista)
    
    for row in lista:
        newRow = []
        for cell in row:
            newCell=Paragraph(str(cell))
            newRow.append(newCell)
        newList.append(newRow)
    
    # Define the table style
    ts = [('ALIGN', (1,0), (-1,-1), 'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LINEABOVE', (0,0), (-1,0), 2, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 2, colors.black),
        ('LINEABOVE', (1,0), (-1,-1), 1, colors.black),
        ('LINEBELOW', (1,0), (-1,-1), 1, colors.black),
        ('FONT', (0,0), (-1,0), 'Times-Bold')]

    # Create the table with its style
    table = Table(newList, style=ts, colWidths=[2*cm, None])
    # Add the table to the Document
    Story.append(table)
    
    return None

def _insert_artifact2(var, Story):
        
    # Transform the dataframe to a list in order to be printed in the document
    lista = [var.columns[:,].values.astype(str).tolist()] + var.values.tolist()
    
    newList = []
    
    for row in lista:
        newRow = []
        for cell in row:
            newCell=Paragraph(str(cell))
            newRow.append(newCell)
        newList.append(newRow)
    
    # Define the table style
    ts = [('ALIGN', (1,0), (-1,-1), 'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LINEABOVE', (0,0), (-1,0), 2, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 2, colors.black),
        ('LINEABOVE', (0,1), (-1,-1), 1, colors.black),
        ('LINEBELOW', (0,1), (-1,-1), 1, colors.black),
        ('FONT', (0,0), (-1,0), 'Times-Bold')]

    # Create the table with its style
    table = Table(newList, style=ts, colWidths=[1.2*cm, 4*cm, 5*cm, 6*cm, 2*cm])
    # Add the table to the Document
    Story.append(table)
    
    return None

def _insert_compound_artifact(keyArtifact, path, sheet, colRange, skipRows, path2, sheet2, colRange2, skipRows2, Story):
    print("Entra al metodo")
    print("Path 1: % s" % (path))
    print("Sheet 1: % s" % (sheet))
    print("Column Range 1: % s" % (colRange))
    print("Skip this row number 1: % s" % (skipRows))
    
    print("Path 2: % s" % (path2))
    print("Sheet 2: % s" % (sheet2))
    print("Column Range 2: % s" % (colRange2))
    print("Skip this row number 2: % s" % (skipRows2))
    
    # To open Workbook
    wb = xl.load_workbook(filename = '.\Artifacts\%s' % (path))
    # Get the active worksheet
    ws = wb.active 
    # Print the artifact's name
    print("The Artifact Title is: %s" % (ws['A3'].value))
    
    # Save the data into a DataFrame from Pandas
    df1 = pd.read_excel('.\Artifacts\%s' % (path), sheet, skiprows = skipRows, header=0, usecols=colRange)
    df2 = pd.read_excel('.\Artifacts\%s' % (path2), sheet2, skiprows = skipRows2, header=0, usecols=colRange2)
    
    #df = df1.set_index(keyArtifact).join(df2.set_index(keyArtifact), lsuffix='_caller')
    #df = df1.set_index(keyArtifact).join(df2, lsuffix='_caller', rsuffix='_other')
    df = df1.join(df2.set_index(keyArtifact), on=keyArtifact, lsuffix='_caller')
    
    # Transform the dataframe to a list in order to be printed in the document
    lista = [df.columns[:,].values.astype(str).tolist()] + df.values.tolist()
    
    newList = []
        
    for row in lista:
        newRow = []
        for cell in row:
            newCell=Paragraph(str(cell))
            newRow.append(newCell)
        newList.append(newRow)
    
    # Define the table style
    ts = [('ALIGN', (1,0), (-1,-1), 'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LINEABOVE', (0,0), (-1,0), 2, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 2, colors.black),
        ('LINEABOVE', (0,1), (-1,-1), 1, colors.black),
        ('LINEBELOW', (0,1), (-1,-1), 1, colors.black),
        ('FONT', (0,0), (-1,0), 'Times-Bold')]

    # Create the table with its style
    table = Table(newList, style=ts, colWidths=[None])
    # Add the table to the Document
    Story.append(table)
    
    return None

def _insert_compound_artifact2(keyArtifact, varName, varName2 ,Story):
    
    df = varName.set_index(keyArtifact).join(varName2.set_index(keyArtifact))   
     
    print(df)
    
    # Transform the dataframe to a list in order to be printed in the document
    lista = [df.columns[:,].values.astype(str).tolist()] + df.values.tolist()
    newList = []
    
    print(lista)
    
    for row in lista:
        newRow = []
        for cell in row:
            newCell=Paragraph(str(cell))
            newRow.append(newCell)
        newList.append(newRow)
    
    # Define the table style
    ts = [('ALIGN', (1,0), (-1,-1), 'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LINEABOVE', (0,0), (-1,0), 2, colors.black),
        ('LINEBELOW', (0,0), (-1,0), 2, colors.black),
        ('LINEABOVE', (1,0), (-1,-1), 1, colors.black),
        ('LINEBELOW', (1,0), (-1,-1), 1, colors.black),
        ('FONT', (0,0), (-1,0), 'Times-Bold')]

    # Create the table with its style
    table = Table(newList, style=ts, colWidths=[1.2*cm, 4*cm, 5*cm, 6*cm, 2*cm])
    # Add the table to the Document
    Story.append(table)
    return None

def _insert_image(path, givenWidth, Story):
    extended_path = ".\Artifacts\%s" % (path)
    #can.drawInlineImage(path, inch*.25, inch*.25, PAGE_WIDTH-(.5*inch), (.316*inch))
    
    Story.append(get_image(extended_path, width = int(givenWidth)*cm))
    return None

def _insert_text(textInput):
    
    print("Va a insertar el texto: " + textInput)
    Story.append(Paragraph(textInput))
    
    return None

def _insert_histogram(path, sheet, colRange, skipRows, xAxis, yAxis, Story):
    # Save the data into a DataFrame from Pandas
    df = pd.read_excel('.\Artifacts\%s' % (path), sheet, skiprows = skipRows, header=0, usecols=colRange)
    
    xAxisContent = xAxis.split(", ")
    xContent = []
    for item in xAxisContent:
        xContent.append(item)
    
    yAxisContent = yAxis.split(", ")  
    yContent = []  
    for item in yAxisContent:
        yContent.append(item)
    
    df.plot(x=xAxis, y=yContent, kind="bar", rot=5, fontsize=4)
    plt.savefig('hist.png')
    Story.append(get_image("hist.png", width = 15*cm))
    
    return None

def _insert_piechart(path, sheet, colRange, skipRows, data, label, Story):
    # Save the data into a DataFrame from Pandas
    df = pd.read_excel('.\Artifacts\%s' % (path), sheet, skiprows = skipRows, header=0, usecols=colRange)
    
    print(df)
    
    print(data)
    print(label)
    
    plt.pie(df[data], labels = df[label])
    plt.savefig('pie.png')
    Story.append(get_image("pie.png", width = 15*cm))
    
    return None

def _loop_row(path, sheet, rowRange, col, content, story):
    # To open Workbook
    wb = xl.load_workbook(filename = '.\Artifacts\%s' % (path))
    # Select sheet
    ws = wb[sheet]
    
    #print(content[0])
        
    # Data array
    data = []
    
    iterRange = col+re.search(r'(\d+):(\d+)', rowRange).group(1)+':'+col+re.search(r'(\d+):(\d+)', rowRange).group(2)
        
    for row in ws[iterRange]:
        for cell in row:
            data.append(cell.value)
            parse_string_array(content, story, value = cell.value)
            
def _loop_sheet(path, content, story):
    print("Está en el loop de hoja")    
    
    sheets_dict = pd.read_excel('.\Artifacts\%s' % (path), sheet_name=None)
    
    sheetName = ""
    
    # Data array
    data = []
        
    for name, sheet in sheets_dict.items():
        #sheet['sheet'] = name
        #sheet = sheet.rename(columns=lambda x: x.split('')[-1])
        sheetName=name
        data.append(name)
        print(name)
        parse_string_array(content, story, value = name)
            
def _loop_col(path, sheet, colRange, row, content, story):
    # To open Workbook
    wb = xl.load_workbook(filename = '.\Artifacts\%s' % (path))
    # Select sheet
    ws = wb[sheet]
            
    # Data array
    data = []
        
    #iterRange = re.search(r'([a-zA-Z]+):([a-zA-Z]+)', colRange).group(1)+row+':'+re.search(r'([a-zA-Z]+):([a-zA-Z]+)', colRange).group(2)+row
    iterRange = "A8:C8"
    #print(iterRange)
        
    for col in ws[iterRange]:
        for cell in col:
            data.append(cell.value)
            parse_string_array(content, story, value = cell.value)
        
def _create_chapter(textInput):
    
    print("Va a crear un capítulo con título: " + textInput)
    Story.append(Paragraph(textInput, styles['Heading1']))
    
    return None

def _create_section(textInput):
    
    print("Va a crear una sección con título: " + textInput)
    Story.append(Paragraph(textInput, styles['Heading2']))
    
    return None

def _conditional(first, second, logic, content, story, **kwargs):
    
    operator_dict = {
        '>': operator.gt,
        '>=': operator.ge,
        '==': operator.eq,
        '!=': operator.ne,
        '<=': operator.le,
        '<': operator.lt 
    }
    
    print(type(operator_dict[logic](first, second)))
    
    if operator_dict[logic](first, second):
        print("Se cumplió la condición y el value es: " + kwargs.get('value', 'cell.value'))
        
        parse_string_array(content, story, value = kwargs.get('value', 'cell.value'))
    
    return None

def get_image(path, width=1*cm):
    img = utils.ImageReader(path)
    iw, ih = img.getSize()
    aspect = ih / float(iw)
    return Image(path, width=width, height=(width * aspect))

def addPageNumberRight(canvas, doc):
    """
    Add the page number
    """
    page_num = canvas.getPageNumber()
    text = "Page #%s" % page_num
    canvas.drawRightString(20*cm, 1*cm, text)
    
    # Add logo to top left corner
    imagePath = ".\Parameters\logo.png"
    img = utils.ImageReader(imagePath)
    iw, ih = img.getSize()
    aspect = ih / float(iw)
    imWidth = cm*2.5
    imHeight = imWidth*aspect
    canvas.drawInlineImage(imagePath, cm*1, cm*26.5, width = imWidth, height = imHeight)
    
    # Add Author's name to top right corner
    # Opening JSON file
    f = open('.\Parameters\doc_params.json')
    params = json.load(f)
    canvas.drawRightString(cm*20, cm*26.5, params['documentParameters']['author'])
    
def addPageNumberCenter(canvas, doc):
    """
    Add the page number
    """
    page_num = canvas.getPageNumber()
    text = "Page #%s" % page_num
    canvas.drawString(10*cm, 1*cm, text)
    
    # Add logo to top left corner
    imagePath = ".\Parameters\logo.png"
    img = utils.ImageReader(imagePath)
    iw, ih = img.getSize()
    aspect = ih / float(iw)
    imWidth = cm*2.5
    imHeight = imWidth*aspect
    canvas.drawInlineImage(imagePath, cm*1, cm*26.5, width = imWidth, height = imHeight)    
    
    # Add Author's name to top right corner
    # Opening JSON file
    f = open('.\Parameters\doc_params.json')
    params = json.load(f)
    canvas.drawRightString(cm*20, cm*26.5, params['documentParameters']['author'])
    
def addPageNumberLeft(canvas, doc):
    """
    Add the page number
    """
    page_num = canvas.getPageNumber()
    text = "Page #%s" % page_num
    canvas.drawString(1*cm, 1*cm, text)
    
    # Add logo to top left corner
    imagePath = ".\Parameters\logo.png"
    img = utils.ImageReader(imagePath)
    iw, ih = img.getSize()
    aspect = ih / float(iw)
    imWidth = cm*2.5
    imHeight = imWidth*aspect
    canvas.drawInlineImage(imagePath, cm*1, cm*26.5, width = imWidth, height = imHeight)
    
    # Add Author's name to top right corner
    # Opening JSON file
    f = open('.\Parameters\doc_params.json')
    params = json.load(f)   
    canvas.drawRightString(cm*20, cm*26.5, params['documentParameters']['author'])
    
def afterFlowable(self, flowable): 
    "Registers TOC entries." 
    if flowable.__class__.__name__ == 'Paragraph':            
        text = flowable.getPlainText()            
        style = flowable.style.name 
        if style == 'Heading1':         
            key = 'h1-%s' % self.seq.nextf('heading1')   
            self.canv.bookmarkPage(key)     
            self.notify('TOCEntry', (0, text, self.page))
        elif style == 'Heading2':   
            key = 'h2-%s' % self.seq.nextf('heading2')   
            self.canv.bookmarkPage(key)              
            self.notify('TOCEntry', (1, text, self.page))         

if __name__ == '__main__':
    # Opening JSON file
    f = open('.\Parameters\doc_params.json')
    
    # returns JSON object as
    # a dictionary
    params = json.load(f)
    
    filepath = '.\instructions.txt'
    pdf_file = params['documentParameters']['documentName']
     
    doc = SimpleDocTemplate(pdf_file,pagesize=letter,
                        rightMargin=float(params['documentParameters']['marginRight'])*cm,leftMargin=float(params['documentParameters']['marginLeft'])*cm,
                        topMargin=float(params['documentParameters']['marginTop'])*cm,bottomMargin=float(params['documentParameters']['marginBottom'])*cm)
    
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Justify', alignment=TA_JUSTIFY))
    Story=[]
    
    toc = TableOfContents()
    toc.levelStyles=[ParagraphStyle(name='Heading1', fontSize=14, leading=16),ParagraphStyle(name='Heading2', fontSize=12, leading=14, leftIndent=20)]
    
    Story.append(toc)
    
    data = parse_file(filepath, Story)
    
    if (params['documentParameters']['pageNum'] == "Right"):
        doc.multiBuild(Story, onFirstPage=addPageNumberRight, onLaterPages=addPageNumberRight)
    elif (params['documentParameters']['pageNum'] == 'Center'):
        doc.multiBuild(Story, onFirstPage=addPageNumberCenter, onLaterPages=addPageNumberCenter)
    elif (params['documentParameters']['pageNum'] == 'Left'):
        doc.multiBuild(Story, onFirstPage=addPageNumberLeft, onLaterPages=addPageNumberLeft)
    else:
        doc.multiBuild(Story)
    
    print(doc.page)
    
    print(data)